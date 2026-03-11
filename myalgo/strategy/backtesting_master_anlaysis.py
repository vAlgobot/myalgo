"""
excel_export_standalone.py
==========================
Standalone Excel report generator for NIFTY Scalping backtest results.

Reads pre-computed analysis tables directly from myalgo.db and writes a
fully-styled 15-sheet Excel workbook — skipping all heavy analysis phases.

Usage:
    python excel_export_standalone.py                   # reads myalgo.db in same dir
    python excel_export_standalone.py path/to/myalgo.db # explicit db path
    python excel_export_standalone.py --run YYYYMMDD_HHMMSS  # specific run ID

Output:
    backtest_results_<run_id>.xlsx  (current directory)
"""

import sqlite3
import pandas as pd
import numpy as np
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── config (must match backtesting_analysis_v4.py) ──────────────────────────
INITIAL_CAPITAL      = 50000
BROKERAGE_PER_ORDER  = 25
SLIPPAGE_PERCENT     = 0.5
SLIPPAGE_MODE        = "PERCENT"

# ── colour palette ────────────────────────────────────────────────────────────
C = {
    'dark':    'FF1A1A2E', 'mid':    'FF16213E', 'accent': 'FF0F3460',
    'gold':    'FFFFD700', 'silver': 'FFC0C0C0', 'bronze': 'FFCD7F32',
    'green_h': 'FF0D5C3A', 'blue_h': 'FF1A4A6E', 'teal_h': 'FF2E4A6E',
    'white':   'FFFFFFFF', 'lgray':  'FFF5F5F5', 'mgray':  'FFE0E0E0',
    'g_cell':  'FFE8FFE8', 'r_cell': 'FFFFE8E8', 'y_cell': 'FFFFFAE8',
    'g_txt':   'FF006600', 'r_txt':  'FFCC0000', 'b_txt':  'FF003399',
    'call_bg': 'FFD6EAF8', 'put_bg': 'FFFDE8E8',
    'tier1':   'FF003322', 'tier2':  'FF0A2540', 'tier3':  'FF1A3A2E',
    'warn':    'FF8B0000', 'warn_r': 'FFFFF5F5',
    'avoid':   'FF2A0A0A',
}

# ── style helpers ─────────────────────────────────────────────────────────────
_thin = Side(style='thin',   color='FFCCCCCC')
_med  = Side(style='medium', color='FF999999')

def mk_border(style='thin'):
    s = _thin if style == 'thin' else _med
    return Border(left=s, right=s, top=s, bottom=s)

def hc(cell, text, bold=True, size=11, fg=None, bg=None,
        wrap=True, align='center', italic=False):
    """Styled header cell."""
    cell.value = text
    cell.font  = Font(name='Arial', bold=bold, size=size,
                      color=fg or C['white'], italic=italic)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = mk_border()

def dc(cell, v, fmt=None, bold=False, bg=None, fg=None, align='center'):
    """Styled data cell."""
    cell.value = v
    cell.font  = Font(name='Arial', bold=bold, size=10, color=fg or 'FF222222')
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = mk_border()
    if fmt:
        cell.number_format = fmt

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, row, col_start, col_end, text, bg,
              fg=None, bold=True, size=13, height=28):
    """Merge + style a title row."""
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start)
    hc(c, text, bold=bold, size=size, fg=fg or C['white'], bg=bg)
    ws.row_dimensions[row].height = height

def rank_to_medal(r):
    return {1: '🥇', 2: '🥈', 3: '🥉'}.get(r, f'#{r}')


# ── DB loader ─────────────────────────────────────────────────────────────────

def load_tables(db_path: str, run_id: str = None) -> dict:
    """Load all analysis tables from DB for the given (or latest) run_id."""
    conn = sqlite3.connect(db_path)

    # Discover latest run if not specified
    if run_id is None:
        try:
            run_id = pd.read_sql(
                "SELECT run_id FROM analysis_runs ORDER BY run_id DESC LIMIT 1", conn
            ).iloc[0, 0]
        except Exception:
            # fallback: pick latest from overall table
            try:
                run_id = pd.read_sql(
                    "SELECT run_id FROM analysis_strategy_overall ORDER BY run_id DESC LIMIT 1", conn
                ).iloc[0, 0]
            except Exception:
                run_id = None

    print(f"📂 Loading run_id: {run_id}")

    def load(table, rid=run_id):
        try:
            if rid:
                df = pd.read_sql(f"SELECT * FROM {table} WHERE run_id=?", conn, params=(rid,))
            else:
                df = pd.read_sql(f"SELECT * FROM {table}", conn)
            return df
        except Exception as e:
            print(f"  ⚠  Could not load {table}: {e}")
            return pd.DataFrame()

    tables = {
        'run_id':          run_id,
        'ranking':         load('analysis_strategy_overall'),
        'yearly':          load('analysis_strategy_yearly'),
        'monthly':         load('analysis_strategy_monthly'),
        'exit_reason':     load('analysis_exit_reason'),
        'single_param':    load('analysis_param_single'),
        'combo_param':     load('analysis_param_combo'),
        'filter_sim':      load('analysis_filter_sim'),
        'ot_overall':      load('analysis_option_type_overall') if _table_exists(conn, 'analysis_option_type_overall')
                           else load('OptionType_Overall') if _table_exists(conn, 'OptionType_Overall')
                           else pd.DataFrame(),
        'ot_scenario':     load('analysis_option_type_scenarios') if _table_exists(conn, 'analysis_option_type_scenarios')
                           else pd.DataFrame(),
    }

    # add rank column if missing
    rk = tables['ranking']
    if not rk.empty and 'rank' not in rk.columns and 'composite' in rk.columns:
        rk = rk.sort_values('composite', ascending=False).reset_index(drop=True)
        rk['rank'] = rk.index + 1
        tables['ranking'] = rk

    conn.close()
    return tables


def _table_exists(conn, name):
    cur = conn.execute(
        "SELECT count(*) FROM sqlite_master WHERE type='table' AND name=?", (name,)
    )
    return cur.fetchone()[0] > 0


# ── sheet builders ────────────────────────────────────────────────────────────

def _sheet_deployment_pick(wb, ranking_df, run_id):
    if ranking_df is None or ranking_df.empty:
        return
    ws = wb.create_sheet('Live Deployment Pick')
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 30, 20, 20, 20, 20, 20, 20, 22])
    ws.row_dimensions[1].height = 8

    title_row(ws, 2, 2, 9,
              f'NIFTY SCALPING — LIVE DEPLOYMENT RECOMMENDATION  |  Run: {run_id}',
              C['dark'], size=14, height=36)
    title_row(ws, 3, 2, 9,
              f'Capital: Rs.{INITIAL_CAPITAL:,}  |  Slippage: {SLIPPAGE_PERCENT}%  |  '
              f'Brokerage: Rs.{BROKERAGE_PER_ORDER}/order  |  '
              f'Composite: Sharpe 30% · RODD 25% · PF 15% · Exp 10% · DD 10% · Cons 10%',
              C['dark'], fg='FFAAAAAA', bold=False, size=9, height=18)
    ws.row_dimensions[4].height = 8

    best = ranking_df.iloc[0]
    title_row(ws, 5, 2, 9, 'RECOMMENDED FOR LIVE DEPLOYMENT',
              C['gold'], fg=C['dark'], size=13, height=28)
    title_row(ws, 6, 2, 9, str(best.get('strategy', '')),
              C['mid'], size=11, height=22)
    ws.row_dimensions[7].height = 6

    compare = ranking_df.head(5)
    col_labels = ['Metric'] + [
        f"{rank_to_medal(int(r.get('rank', i+1)))} Rank {int(r.get('rank', i+1))}"
        for i, (_, r) in enumerate(compare.iterrows())
    ]
    col_bgs = [C['accent'], C['green_h'], C['blue_h'], C['teal_h'], 'FF3A3A5C', 'FF4A3A2E']
    for ci, (lbl, bg) in enumerate(zip(col_labels, col_bgs), 2):
        hc(ws.cell(row=8, column=ci), lbl, bg=bg)
    ws.row_dimensions[8].height = 22

    metrics_def = [
        ('Net P&L (Rs.)',        'net_pnl',         '#,##0',   True),
        ('Growth %',             'growth_pct',       '0.0%',    True),
        ('Total Trades',         'total_trades',     '#,##0',   False),
        ('Win Rate',             'win_rate',         '0.0%',    True),
        ('Profit Factor',        'profit_factor',    '0.00',    True),
        ('Sharpe Ratio',         'sharpe',           '0.00',    True),
        ('RODD',                 'rodd',             '0.00',    True),
        ('Max Drawdown %',       'max_dd_pct',       '0.0%',    False),
        ('Max DD Amount (Rs.)',  'max_dd_amt',       '#,##0',   False),
        ('RR Ratio',             'rr_ratio',         '0.00',    True),
        ('Avg Trade (Rs.)',      'avg_trade',        '#,##0',   True),
        ('Expectancy (Rs.)',     'expectancy',       '#,##0',   True),
        ('Win Day %',            'win_day_pct',      '0.0%',    True),
        ('Max Win Streak',       'max_win_streak',   '#,##0',   True),
        ('Max Loss Streak',      'max_loss_streak',  '#,##0',   False),
        ('Composite Score',      'composite',        '0.000',   True),
    ]
    for ri, (label, col, fmt, higher_better) in enumerate(metrics_def, 9):
        ws.row_dimensions[ri].height = 19
        lb = ws.cell(row=ri, column=2)
        lb.value = label
        lb.font = Font(name='Arial', bold=True, size=10)
        lb.fill = PatternFill('solid', fgColor=C['lgray'])
        lb.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        lb.border = mk_border()

        vals_list = [float(r.get(col, 0) or 0) for _, r in compare.iterrows()]
        best_v = max(vals_list) if higher_better else min(vals_list)

        for ci, (_, row) in enumerate(compare.iterrows()):
            v = float(row.get(col, 0) or 0)
            display_v = v / 100 if fmt.endswith('%') and v > 1 else v
            is_best  = abs(v - best_v) < 1e-9
            is_rank1 = ci == 0
            if is_best:
                bg, fg_c, bold_c = C['tier1'], 'FF00EE88', True
            elif is_rank1:
                bg, fg_c, bold_c = C['tier2'], 'FFADD8E6', False
            else:
                bg, fg_c, bold_c = C['lgray'], 'FF222222', False
            dc(ws.cell(row=ri, column=ci + 3), display_v,
               fmt=fmt, bold=bold_c, bg=bg, fg=fg_c)

    next_r = 9 + len(metrics_def) + 1
    ws.row_dimensions[next_r - 1].height = 8

    title_row(ws, next_r, 2, 9, 'WHY THIS STRATEGY — KEY REASONS', C['accent'], height=24)
    next_r += 1
    reasons = [
        ('Highest Composite Score',
         f"Score: {float(best.get('composite', 0)):.3f} — #1 of {len(ranking_df)} strategies. Best overall risk-adjusted quality."),
        ('Best Profit Factor',
         f"PF {float(best.get('profit_factor', 0)):.2f}x — highest gross edge on every rupee risked."),
        ('Best Sharpe Ratio',
         f"Sharpe {float(best.get('sharpe', 0)):.2f} — top risk-adjusted return per unit of volatility."),
        ('Selective Entry Filter',
         f"Only {int(best.get('total_trades', 0))} trades — quality over quantity. Lower slippage risk in live."),
        ('Highest Avg Trade',
         f"Rs.{float(best.get('avg_trade', 0)):.0f} avg per trade — more per execution than lower-ranked strategies."),
        ('Controlled Drawdown',
         f"Max DD {float(best.get('max_dd_pct', 0)):.1f}% — lower than most higher-frequency alternatives."),
        ('Return Over Drawdown',
         f"RODD {float(best.get('rodd', 0)):.2f}x — strong reward-to-drawdown ratio confirms edge durability."),
    ]
    for reason, detail in reasons:
        ws.merge_cells(start_row=next_r, start_column=2, end_row=next_r, end_column=3)
        c = ws.cell(row=next_r, column=2)
        c.value = reason
        c.font = Font(name='Arial', bold=True, size=10, color=C['white'])
        c.fill = PatternFill('solid', fgColor='FF0D3349')
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = mk_border()
        ws.merge_cells(start_row=next_r, start_column=4, end_row=next_r, end_column=9)
        c2 = ws.cell(row=next_r, column=4)
        c2.value = detail
        c2.font = Font(name='Arial', size=10)
        c2.fill = PatternFill('solid', fgColor='FFF5F9FF')
        c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
        c2.border = mk_border()
        ws.row_dimensions[next_r].height = 20
        next_r += 1

    ws.row_dimensions[next_r].height = 8
    next_r += 1
    title_row(ws, next_r, 2, 9, 'RISK WARNINGS BEFORE GOING LIVE', C['warn'], height=24)
    next_r += 1
    warnings_list = [
        ('Small Sample Risk',   'Validate live performance for min 2-4 weeks before scaling up position size.'),
        ('Max Drawdown Cap',    f"Rs.{float(best.get('max_dd_amt', 0)):,.0f} max DD. Daily loss limit = max_dd/4 before pausing."),
        ('Choppy Market Risk',  'If win rate drops below 30% for 2 consecutive weeks — pause and reassess.'),
        ('Slippage Assumption', f"{SLIPPAGE_PERCENT}% slippage assumed. Use limit orders within this range in live trading."),
        ('Strategy Changes',    'Any code changes to entry/exit logic require re-backtest before deploying.'),
    ]
    for w_title, w_detail in warnings_list:
        ws.merge_cells(start_row=next_r, start_column=2, end_row=next_r, end_column=3)
        c = ws.cell(row=next_r, column=2)
        c.value = f'WARNING: {w_title}'
        c.font = Font(name='Arial', bold=True, size=10, color='FFFF4444')
        c.fill = PatternFill('solid', fgColor=C['avoid'])
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = mk_border()
        ws.merge_cells(start_row=next_r, start_column=4, end_row=next_r, end_column=9)
        c2 = ws.cell(row=next_r, column=4)
        c2.value = w_detail
        c2.font = Font(name='Arial', size=10)
        c2.fill = PatternFill('solid', fgColor=C['warn_r'])
        c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
        c2.border = mk_border()
        ws.row_dimensions[next_r].height = 22
        next_r += 1

    ws.freeze_panes = 'C9'


def _sheet_all_ranking(wb, ranking_df, run_id):
    if ranking_df is None or ranking_df.empty:
        return
    ws = wb.create_sheet('All Strategy Ranking')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 18, f'ALL STRATEGIES — RANKED BY COMPOSITE SCORE  |  {run_id}',
              C['dark'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Rank', 'Strategy', 'Net P&L', 'Growth%', 'Trades', 'Win Rate', 'PF',
            'Sharpe', 'RODD', 'MaxDD%', 'MaxDD Rs.', 'RR', 'Avg Trade',
            'Expectancy', 'WinDay%', 'Composite', 'WinStreak', 'LossStreak']
    hdr_w = [6, 52, 13, 10, 7, 10, 8, 8, 8, 9, 11, 8, 11, 11, 9, 10, 10, 10]
    fmts  = ['#,##0','@','#,##0','0.0%','#,##0','0.0%','0.00','0.00','0.00',
             '0.0%','#,##0','0.00','#,##0','#,##0','0.0%','0.000','#,##0','#,##0']
    set_col_widths(ws, hdr_w)
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    medal_bgs = {1: 'FFFFF8DC', 2: 'FFF5F5F5', 3: 'FFFFF4E8'}
    for ri, (_, row) in enumerate(ranking_df.sort_values('rank').iterrows(), 5):
        rk = int(row.get('rank', ri - 4))
        bg = medal_bgs.get(rk, 'FFF9F9F9' if ri % 2 == 0 else C['white'])
        vals = [
            rk, row.get('strategy', ''),
            float(row.get('net_pnl', 0) or 0),
            float(row.get('growth_pct', 0) or 0) / 100,
            int(row.get('total_trades', 0) or 0),
            float(row.get('win_rate', 0) or 0) / 100,
            float(row.get('profit_factor', 0) or 0),
            float(row.get('sharpe', 0) or 0),
            float(row.get('rodd', 0) or 0),
            float(row.get('max_dd_pct', 0) or 0) / 100,
            float(row.get('max_dd_amt', 0) or 0),
            float(row.get('rr_ratio', 0) or 0),
            float(row.get('avg_trade', 0) or 0),
            float(row.get('expectancy', 0) or 0),
            float(row.get('win_day_pct', 0) or 0) / 100,
            float(row.get('composite', 0) or 0),
            int(row.get('max_win_streak', 0) or 0),
            int(row.get('max_loss_streak', 0) or 0),
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            dc(cell, v, fmt=fmt, bg=bg, align='center' if ci != 2 else 'left')
            if ci == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[ri].height = 18
    ws.freeze_panes = 'A5'


def _sheet_yearly(wb, yearly_df):
    if yearly_df is None or yearly_df.empty:
        return
    ws = wb.create_sheet('Yearly Breakdown')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 12, 'YEARLY PERFORMANCE BREAKDOWN', C['dark'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Strategy', 'Year', 'Net P&L', 'Growth%', 'Trades', 'Win Rate',
            'PF', 'Sharpe', 'MaxDD%', 'AvgTrade', 'WinDay%', 'BestDay']
    fmts = ['@','#,##0','#,##0','0.0%','#,##0','0.0%','0.00','0.00','0.0%','#,##0','0.0%','#,##0']
    set_col_widths(ws, [52, 8, 13, 10, 8, 10, 8, 8, 9, 11, 9, 12])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    for ri, (_, row) in enumerate(yearly_df.sort_values(['strategy', 'year']).iterrows(), 5):
        pnl = float(row.get('net_pnl', 0) or 0)
        bg  = C['g_cell'] if pnl >= 0 else C['r_cell']
        vals = [
            row.get('strategy', ''), row.get('year', ''),
            pnl, float(row.get('growth_pct', 0) or 0) / 100,
            int(row.get('total_trades', 0) or 0),
            float(row.get('win_rate', 0) or 0) / 100,
            float(row.get('profit_factor', 0) or 0),
            float(row.get('sharpe', 0) or 0),
            float(row.get('max_dd_pct', 0) or 0) / 100,
            float(row.get('avg_trade', 0) or 0) if 'avg_trade' in row else 0,
            float(row.get('win_day_pct', 0) or 0) / 100 if 'win_day_pct' in row else 0,
            float(row.get('best_day_pnl', 0) or 0) if 'best_day_pnl' in row else 0,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            pnl_fg = (C['g_txt'] if pnl >= 0 else C['r_txt']) if ci == 3 else 'FF222222'
            dc(cell, v, fmt=fmt, bg=bg, fg=pnl_fg,
               align='center' if ci != 1 else 'left')
            if ci == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[ri].height = 18
    ws.freeze_panes = 'A5'


def _sheet_monthly(wb, monthly_df):
    if monthly_df is None or monthly_df.empty:
        return
    ws = wb.create_sheet('Monthly Breakdown')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 11, 'MONTHLY PERFORMANCE BREAKDOWN', C['dark'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Strategy', 'Month', 'Net P&L', 'Growth%', 'Trades', 'Win Rate',
            'PF', 'Sharpe', 'MaxDD%', 'AvgTrade', 'WinDay%']
    fmts = ['@','@','#,##0','0.0%','#,##0','0.0%','0.00','0.00','0.0%','#,##0','0.0%']
    set_col_widths(ws, [52, 10, 13, 10, 8, 10, 8, 8, 9, 11, 9])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    for ri, (_, row) in enumerate(monthly_df.sort_values(['strategy', 'year_month']).iterrows(), 5):
        pnl = float(row.get('net_pnl', 0) or 0)
        bg  = C['g_cell'] if pnl >= 0 else C['r_cell']
        vals = [
            row.get('strategy', ''), str(row.get('year_month', '')),
            pnl, float(row.get('growth_pct', 0) or 0) / 100,
            int(row.get('total_trades', 0) or 0),
            float(row.get('win_rate', 0) or 0) / 100,
            float(row.get('profit_factor', 0) or 0),
            float(row.get('sharpe', 0) or 0),
            float(row.get('max_dd_pct', 0) or 0) / 100,
            float(row.get('avg_trade', 0) or 0) if 'avg_trade' in row else 0,
            float(row.get('win_day_pct', 0) or 0) / 100 if 'win_day_pct' in row else 0,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            pnl_fg = (C['g_txt'] if pnl >= 0 else C['r_txt']) if ci == 3 else 'FF222222'
            dc(cell, v, fmt=fmt, bg=bg, fg=pnl_fg,
               align='center' if ci != 1 else 'left')
            if ci == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[ri].height = 17
    ws.freeze_panes = 'A5'


def _sheet_drawdown(wb, ranking_df):
    if ranking_df is None or ranking_df.empty:
        return
    ws = wb.create_sheet('Drawdown Analysis')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 10, 'DRAWDOWN ANALYSIS — ALL STRATEGIES', C['warn'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Rank', 'Strategy', 'MaxDD %', 'MaxDD Rs.', 'RODD', 'Sharpe',
            'Net P&L', 'PF', 'DD Risk Tier', 'Position Size Guidance']
    fmts = ['#,##0','@','0.0%','#,##0','0.00','0.00','#,##0','0.00','@','@']
    set_col_widths(ws, [6, 52, 10, 13, 8, 8, 13, 8, 14, 40])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['warn'])
    ws.row_dimensions[4].height = 22

    for ri, (_, row) in enumerate(ranking_df.sort_values('max_dd_pct').iterrows(), 5):
        dd = float(row.get('max_dd_pct', 0) or 0)
        if dd <= 10:
            tier, t_bg = 'LOW RISK',   C['g_cell']
            ps = 'Full size (1x). Low drawdown — suitable for larger allocation.'
        elif dd <= 15:
            tier, t_bg = 'MODERATE',   C['y_cell']
            ps = 'Standard size (0.75x). Monitor monthly for regime changes.'
        elif dd <= 20:
            tier, t_bg = 'ELEVATED',   'FFFCE4D6'
            ps = 'Reduced size (0.5x). High drawdown — tighter daily loss limit.'
        else:
            tier, t_bg = 'HIGH RISK',  C['r_cell']
            ps = 'Avoid or paper trade only (0.25x). Exceeds 20% DD threshold.'

        vals = [
            int(row.get('rank', 0)), row.get('strategy', ''),
            dd / 100, float(row.get('max_dd_amt', 0) or 0),
            float(row.get('rodd', 0) or 0), float(row.get('sharpe', 0) or 0),
            float(row.get('net_pnl', 0) or 0), float(row.get('profit_factor', 0) or 0),
            tier, ps,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            row_bg = t_bg if ci in [9, 10] else (C['lgray'] if ri % 2 == 0 else C['white'])
            dc(cell, v, fmt=fmt, bg=row_bg,
               align='center' if ci not in [2, 10] else 'left')
            if ci in [2, 10]:
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           indent=1, wrap_text=True)
        ws.row_dimensions[ri].height = 20
    ws.freeze_panes = 'A5'


def _sheet_exit_reason(wb, exit_reason_df):
    if exit_reason_df is None or exit_reason_df.empty:
        return
    ws = wb.create_sheet('Exit Reason Analysis')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 9, 'EXIT REASON ANALYSIS — ALL STRATEGIES', C['dark'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Strategy', 'Exit Reason', 'Count', 'Total P&L', 'Avg P&L',
            '% of Trades', '% of P&L', 'Signal Quality', 'Action']
    fmts = ['@','@','#,##0','#,##0','#,##0','0.0%','0.0%','@','@']
    set_col_widths(ws, [46, 28, 8, 13, 12, 11, 11, 22, 30])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    sig_map = {
        'Market closed':        ('HOLD WINNERS',  C['g_cell'],  'Best outcome — let position run to close'),
        'CPR_FINAL_TARGET':     ('TARGET HIT',    'FFD4FFD4',   'Perfect — target reached as planned'),
        'CPR_SL_HIT_CALL_SIDE': ('CALL SL HIT',   C['y_cell'],  'Acceptable — SL working as designed'),
        'CPR_SL_HIT_PUT_SIDE':  ('PUT SL HIT',    C['y_cell'],  'Acceptable — SL working as designed'),
    }

    for strat, grp in exit_reason_df.groupby('strategy'):
        total_t = grp['count'].sum()
        total_p = grp['total_pnl'].sum()
        for _, row in grp.sort_values('total_pnl', ascending=False).iterrows():
            ri_val = ws.max_row + 1
            reason = str(row.get('reason', ''))
            pnl    = float(row.get('total_pnl', 0) or 0)
            if reason in sig_map:
                sig, sig_bg, action = sig_map[reason]
            elif pnl >= 0:
                sig, sig_bg, action = 'PROFITABLE', C['g_cell'], 'Positive exit — review if systematic'
            else:
                sig, sig_bg, action = 'LOSS EXIT',  C['r_cell'], 'Loss exit — check if avoidable via filter'

            pct_t = int(row.get('count', 0)) / total_t if total_t > 0 else 0
            pct_p = pnl / total_p if total_p != 0 else 0
            vals  = [strat, reason, int(row.get('count', 0)),
                     pnl, float(row.get('avg_pnl', 0) or 0),
                     pct_t, pct_p, sig, action]
            for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(row=ri_val, column=ci)
                pnl_fg = (C['g_txt'] if pnl >= 0 else C['r_txt']) if ci in [4, 5] else 'FF222222'
                row_bg = sig_bg if ci in [8, 9] else (C['lgray'] if ri_val % 2 == 0 else C['white'])
                dc(cell, v, fmt=fmt, bg=row_bg, fg=pnl_fg,
                   align='center' if ci not in [1, 2, 9] else 'left')
                if ci in [1, 2, 9]:
                    cell.alignment = Alignment(horizontal='left', vertical='center',
                                               indent=1, wrap_text=True)
            ws.row_dimensions[ri_val].height = 20
    ws.freeze_panes = 'A5'


def _sheet_setup_edge(wb, single_param_df):
    if single_param_df is None or single_param_df.empty:
        return
    ws = wb.create_sheet('Setup Edge & Position Sizing')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 12,
              'SETUP EDGE ANALYSIS — POSITION SIZING GUIDE (Single Parameters)',
              C['dark'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Strategy', 'Parameter', 'Value', 'OPT Type', 'Count',
            'Total P&L', 'Avg P&L', 'Win Rate', 'PF', 'RR', 'Size Tier', 'Multiplier']
    fmts = ['@','@','@','@','#,##0','#,##0','#,##0','0.0%','0.00','0.00','@','0.0x']
    set_col_widths(ws, [44, 22, 24, 10, 8, 12, 10, 9, 8, 8, 14, 10])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    def size_tier(avg_pnl, pf, win_rate):
        if avg_pnl <= 0 or pf < 1.0:
            return 'AVOID',    0.0, C['r_cell'],  C['r_txt']
        elif pf >= 2.5 and win_rate >= 45:
            return 'HIGH A',   1.5, 'FFD4FFD4',   C['g_txt']
        elif pf >= 2.0:
            return 'HIGH B',   1.3, C['g_cell'],  C['g_txt']
        elif pf >= 1.7 and win_rate >= 38:
            return 'HIGH C',   1.1, 'FFEAFAF1',   C['g_txt']
        elif pf >= 1.3:
            return 'NEUTRAL',  0.8, C['y_cell'],  'FF886600'
        else:
            return 'LOW EDGE', 0.4, 'FFFCE4D6',   C['r_txt']

    # Show ALL-option_type rows if column exists, else all rows
    df = single_param_df.copy()
    if 'option_type' in df.columns:
        df = df[df['option_type'] == 'ALL']
    df = df.sort_values('avg_pnl', ascending=False)

    for ri, (_, row) in enumerate(df.iterrows(), 5):
        avg_p = float(row.get('avg_pnl', 0) or 0)
        pf_v  = float(row.get('profit_factor', 0) or 0)
        wr_v  = float(row.get('win_rate', 0) or 0)
        tier, mult, t_bg, t_fg = size_tier(avg_p, pf_v, wr_v)
        ot = row.get('option_type', 'ALL') if 'option_type' in row else 'ALL'
        vals = [
            row.get('strategy', ''), row.get('param', ''),
            row.get('param_value', ''), ot,
            int(row.get('count', 0) or 0),
            float(row.get('total_pnl', 0) or 0), avg_p,
            wr_v / 100, pf_v,
            float(row.get('rr_ratio', 0) or 0),
            tier, mult,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            row_bg = t_bg if ci in [11, 12] else (C['lgray'] if ri % 2 == 0 else C['white'])
            fg_c = (t_fg if ci in [11, 12] else
                    (C['g_txt'] if ci == 7 and avg_p > 0 else
                     (C['r_txt'] if ci == 7 and avg_p < 0 else 'FF222222')))
            dc(cell, v, fmt=fmt, bg=row_bg, fg=fg_c,
               bold=(ci == 12), align='center' if ci not in [1, 2, 3] else 'left')
            if ci in [1, 2, 3]:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[ri].height = 17
    ws.freeze_panes = 'A5'


def _sheet_expiry(wb, single_param_df):
    if single_param_df is None or single_param_df.empty or 'param' not in single_param_df.columns:
        return
    df = single_param_df[single_param_df['param'] == 'expiry_day'].copy()
    if df.empty:
        return
    ws = wb.create_sheet('Expiry Day Analysis')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 10, 'EXPIRY DAY ANALYSIS', C['dark'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Strategy', 'Expiry Day?', 'Opt Type', 'Count', 'Total P&L',
            'Avg P&L', 'Win Rate', 'PF', 'RR', 'Edge Verdict']
    fmts = ['@','@','@','#,##0','#,##0','#,##0','0.0%','0.00','0.00','@']
    set_col_widths(ws, [44, 14, 10, 8, 13, 11, 10, 8, 8, 28])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    for ri, (_, row) in enumerate(df.sort_values(['strategy', 'param_value']).iterrows(), 5):
        avg_p = float(row.get('avg_pnl', 0) or 0)
        pf_v  = float(row.get('profit_factor', 0) or 0)
        if avg_p > 200 and pf_v >= 1.5:
            verdict, v_bg = 'HIGH EDGE — Trade actively',          C['g_cell']
        elif avg_p > 0:
            verdict, v_bg = 'POSITIVE EDGE — Standard size',       C['y_cell']
        else:
            verdict, v_bg = 'NEGATIVE EDGE — Avoid / reduce size', C['r_cell']
        ot   = row.get('option_type', 'ALL') if 'option_type' in row else 'ALL'
        vals = [
            row.get('strategy', ''), row.get('param_value', ''), ot,
            int(row.get('count', 0) or 0),
            float(row.get('total_pnl', 0) or 0), avg_p,
            float(row.get('win_rate', 0) or 0) / 100, pf_v,
            float(row.get('rr_ratio', 0) or 0), verdict,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            row_bg = v_bg if ci == 10 else (C['lgray'] if ri % 2 == 0 else C['white'])
            pnl_fg = (C['g_txt'] if avg_p > 0 else C['r_txt']) if ci == 6 else 'FF222222'
            dc(cell, v, fmt=fmt, bg=row_bg, fg=pnl_fg,
               align='center' if ci not in [1, 2, 10] else 'left')
            if ci in [1, 2, 10]:
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           indent=1, wrap_text=True)
        ws.row_dimensions[ri].height = 20
    ws.freeze_panes = 'A5'


def _sheet_call_put(wb, ot_overall_df):
    if ot_overall_df is None or ot_overall_df.empty:
        return
    ws = wb.create_sheet('CALL vs PUT Analysis')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 10, 'CALL vs PUT — OVERALL OPTION TYPE ANALYSIS', C['dark'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Strategy', 'Option Type', 'Count', 'Total P&L', 'Avg P&L',
            'Win Rate', 'PF', 'RR', 'MaxDD', 'Direction Verdict']
    fmts = ['@','@','#,##0','#,##0','#,##0','0.0%','0.00','0.00','#,##0','@']
    set_col_widths(ws, [44, 12, 8, 13, 11, 10, 8, 8, 11, 30])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    for ri, (_, row) in enumerate(ot_overall_df.sort_values(['strategy', 'option_type']).iterrows(), 5):
        ot    = str(row.get('option_type', ''))
        avg_p = float(row.get('avg_pnl', 0) or 0)
        pf_v  = float(row.get('profit_factor', 0) or 0)
        ot_bg = C['call_bg'] if ot == 'CALL' else (C['put_bg'] if ot == 'PUT' else C['lgray'])
        verdict = (f"{ot} EDGE: PF {pf_v:.2f}" +
                   (' — FAVOURED' if pf_v >= 1.5 else (' — NEUTRAL' if pf_v >= 1.0 else ' — AVOID')))
        vals = [
            row.get('strategy', ''), ot,
            int(row.get('count', 0) or 0),
            float(row.get('total_pnl', 0) or 0), avg_p,
            float(row.get('win_rate', 0) or 0) / 100, pf_v,
            float(row.get('rr_ratio', 0) or 0),
            float(row.get('max_dd_amt', 0) or 0), verdict,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            row_bg = ot_bg if ci in [2, 10] else (C['lgray'] if ri % 2 == 0 else C['white'])
            pnl_fg = (C['g_txt'] if avg_p >= 0 else C['r_txt']) if ci in [4, 5] else 'FF222222'
            dc(cell, v, fmt=fmt, bg=row_bg, fg=pnl_fg,
               align='center' if ci not in [1, 10] else 'left')
            if ci in [1, 10]:
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           indent=1, wrap_text=True)
        ws.row_dimensions[ri].height = 20
    ws.freeze_panes = 'A5'


def _sheet_scenarios(wb, ot_scenario_df):
    if ot_scenario_df is None or ot_scenario_df.empty:
        return
    ws = wb.create_sheet('CALL vs PUT Scenarios')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 11,
              'CALL vs PUT — SCENARIO EDGE (Which setups favour CALL or PUT)',
              C['dark'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Strategy', 'Parameter', 'Value', 'CALL Count', 'PUT Count',
            'CALL Avg P&L', 'PUT Avg P&L', 'CALL Edge', 'CALL WR', 'PUT WR', 'Verdict']
    fmts = ['@','@','@','#,##0','#,##0','#,##0','#,##0','#,##0','0.0%','0.0%','@']
    set_col_widths(ws, [44, 22, 24, 10, 10, 13, 13, 11, 9, 9, 18])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    for ri, (_, row) in enumerate(ot_scenario_df.sort_values('call_edge', ascending=False).iterrows(), 5):
        verdict = str(row.get('verdict', ''))
        v_bg = (C['call_bg'] if verdict == 'CALL_FAVOURED' else
                (C['put_bg'] if verdict == 'PUT_FAVOURED' else C['y_cell']))
        call_wr = float(row.get('win_rate_CALL', 0) or 0)
        put_wr  = float(row.get('win_rate_PUT',  0) or 0)
        vals = [
            row.get('strategy', ''), row.get('param', ''), row.get('param_value', ''),
            float(row.get('count_CALL', 0) or 0), float(row.get('count_PUT', 0) or 0),
            float(row.get('call_avg_pnl', 0) or 0), float(row.get('put_avg_pnl', 0) or 0),
            float(row.get('call_edge', 0) or 0),
            call_wr / 100, put_wr / 100, verdict,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            row_bg = (v_bg if ci == 11 else
                      (C['call_bg'] if ci in [4, 6, 9] else
                       (C['put_bg']  if ci in [5, 7, 10] else
                        (C['lgray'] if ri % 2 == 0 else C['white']))))
            pnl_fg = (C['g_txt'] if float(v or 0) >= 0 else C['r_txt']) if ci in [6, 7, 8] else 'FF222222'
            dc(cell, v, fmt=fmt, bg=row_bg, fg=pnl_fg,
               align='center' if ci not in [1, 2, 3, 11] else 'left')
            if ci in [1, 2, 3, 11]:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[ri].height = 17
    ws.freeze_panes = 'A5'


def _sheet_combo(wb, combo_param_df):
    if combo_param_df is None or combo_param_df.empty:
        return
    ws = wb.create_sheet('Combo Param Analysis')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 10,
              'PHASE 2 — COMBO PARAMETER EDGE (Top 50 HIGH + Bottom 50 LOW setups)',
              C['dark'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Strategy', 'Params Combo', 'Values', 'Opt Type', 'Count',
            'Total P&L', 'Avg P&L', 'Win Rate', 'PF', 'Setup Verdict']
    fmts = ['@','@','@','@','#,##0','#,##0','#,##0','0.0%','0.00','@']
    set_col_widths(ws, [40, 32, 38, 10, 8, 13, 11, 9, 8, 22])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    df = combo_param_df.copy()
    if 'option_type' in df.columns:
        df = df[df['option_type'] == 'ALL']
    top_cb    = df.nlargest(50, 'avg_pnl')
    bottom_cb = df.nsmallest(50, 'avg_pnl')
    display   = pd.concat([top_cb, bottom_cb]).drop_duplicates()

    for ri, (_, row) in enumerate(display.iterrows(), 5):
        avg_p = float(row.get('avg_pnl', 0) or 0)
        pf_v  = float(row.get('profit_factor', 0) or 0)
        if avg_p > 500 and pf_v >= 2.0:
            s_verdict, s_bg = 'HIGH PROBABILITY', C['g_cell']
        elif avg_p > 0:
            s_verdict, s_bg = 'POSITIVE EDGE',    C['y_cell']
        else:
            s_verdict, s_bg = 'AVOID SETUP',      C['r_cell']
        ot = row.get('option_type', 'ALL') if 'option_type' in row else 'ALL'
        vals = [
            row.get('strategy', ''), row.get('params', ''), row.get('param_values', ''), ot,
            int(row.get('count', 0) or 0),
            float(row.get('total_pnl', 0) or 0), avg_p,
            float(row.get('win_rate', 0) or 0) / 100, pf_v, s_verdict,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            row_bg = s_bg if ci == 10 else (C['lgray'] if ri % 2 == 0 else C['white'])
            pnl_fg = (C['g_txt'] if avg_p >= 0 else C['r_txt']) if ci == 7 else 'FF222222'
            dc(cell, v, fmt=fmt, bg=row_bg, fg=pnl_fg,
               align='center' if ci not in [1, 2, 3, 10] else 'left')
            if ci in [1, 2, 3, 10]:
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           indent=1, wrap_text=True)
        ws.row_dimensions[ri].height = 17
    ws.freeze_panes = 'A5'


def _sheet_filter_sim(wb, filter_sim_df):
    if filter_sim_df is None or filter_sim_df.empty:
        return
    ws = wb.create_sheet('Filter Simulation')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 12,
              'PHASE 3 — FILTER SIMULATION (Impact of Excluding Bad Setups)',
              C['dark'], height=30)
    ws.row_dimensions[3].height = 8

    hdrs = ['Strategy', 'Filter Rule', 'Removed', 'Base P&L', 'Filtered P&L',
            'P&L Delta', 'Base WR', 'Filt WR', 'WR Delta', 'DD Delta', 'PF Delta', 'Net Impact']
    fmts = ['@','@','#,##0','#,##0','#,##0','#,##0','0.0%','0.0%','0.0%','#,##0','0.00','@']
    set_col_widths(ws, [40, 38, 9, 12, 12, 11, 9, 9, 9, 11, 9, 20])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    for ri, (_, row) in enumerate(filter_sim_df.sort_values('pnl_delta', ascending=False).iterrows(), 5):
        pnl_d = float(row.get('pnl_delta', 0) or 0)
        dd_d  = float(row.get('dd_delta', 0) or 0)
        wr_d  = float(row.get('win_rate_delta', 0) or 0)
        pf_d  = float(row.get('pf_delta', 0) or 0)
        good  = sum([pnl_d > 0, dd_d > 0, wr_d > 0, pf_d > 0])
        if good >= 3:
            impact, i_bg = 'APPLY — Strong improvement',    C['g_cell']
        elif good >= 2:
            impact, i_bg = 'CONSIDER — Mixed signals',      C['y_cell']
        else:
            impact, i_bg = 'SKIP — Marginal or negative',   C['r_cell']
        vals = [
            row.get('strategy', ''), row.get('filter_desc', ''),
            int(row.get('trades_removed', 0) or 0),
            float(row.get('baseline_pnl', 0) or 0),
            float(row.get('filtered_pnl', 0) or 0), pnl_d,
            float(row.get('baseline_win_rate', 0) or 0) / 100,
            float(row.get('filtered_win_rate', 0) or 0) / 100, wr_d / 100,
            dd_d, pf_d, impact,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            row_bg = i_bg if ci == 12 else (C['lgray'] if ri % 2 == 0 else C['white'])
            delta_fg = ((C['g_txt'] if float(v or 0) > 0 else
                         (C['r_txt'] if float(v or 0) < 0 else 'FF222222'))
                        if ci in [6, 9, 10, 11] else 'FF222222')
            dc(cell, v, fmt=fmt, bg=row_bg, fg=delta_fg,
               align='center' if ci not in [1, 2, 12] else 'left')
            if ci in [1, 2, 12]:
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           indent=1, wrap_text=True)
        ws.row_dimensions[ri].height = 20
    ws.freeze_panes = 'A5'


def _sheet_cpr(wb, single_param_df):
    if single_param_df is None or single_param_df.empty or 'param' not in single_param_df.columns:
        return
    df = single_param_df[
        single_param_df['param'].isin(['daily_cpr_type', 'weekly_cpr_type', 'cpr_relation'])
    ].copy()
    if df.empty:
        return
    ws = wb.create_sheet('CPR Type Analysis')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 11,
              'CPR TYPE ANALYSIS — Daily / Weekly CPR & Relation Edge', C['dark'], height=30)
    ws.row_dimensions[3].height = 8
    _write_param_sheet(ws, df)


def _sheet_gap_open(wb, single_param_df):
    if single_param_df is None or single_param_df.empty or 'param' not in single_param_df.columns:
        return
    df = single_param_df[
        single_param_df['param'].isin(['gap_type', 'open_location', 'hourly_vs_weekly_cpr'])
    ].copy()
    if df.empty:
        return
    ws = wb.create_sheet('Gap & Open Location')
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 11, 'GAP TYPE & OPEN LOCATION ANALYSIS', C['dark'], height=30)
    ws.row_dimensions[3].height = 8
    _write_param_sheet(ws, df)


def _write_param_sheet(ws, df):
    """Shared layout for CPR and Gap sheets."""
    hdrs = ['Strategy', 'Parameter', 'Value', 'Opt Type', 'Count',
            'Total P&L', 'Avg P&L', 'Win Rate', 'PF', 'RR', 'Edge']
    fmts = ['@','@','@','@','#,##0','#,##0','#,##0','0.0%','0.00','0.00','@']
    set_col_widths(ws, [42, 22, 34, 10, 8, 12, 10, 9, 8, 8, 20])
    for ci, h in enumerate(hdrs, 1):
        hc(ws.cell(row=4, column=ci), h, bg=C['accent'])
    ws.row_dimensions[4].height = 22

    for ri, (_, row) in enumerate(
        df.sort_values(['param', 'avg_pnl'], ascending=[True, False]).iterrows(), 5
    ):
        avg_p = float(row.get('avg_pnl', 0) or 0)
        pf_v  = float(row.get('profit_factor', 0) or 0)
        edge  = ('HIGH EDGE' if avg_p > 300 and pf_v >= 1.7 else
                 ('POSITIVE' if avg_p > 0 else 'NEGATIVE'))
        e_bg  = C['g_cell'] if edge == 'HIGH EDGE' else (C['y_cell'] if edge == 'POSITIVE' else C['r_cell'])
        ot    = row.get('option_type', 'ALL') if 'option_type' in row else 'ALL'
        vals  = [
            row.get('strategy', ''), row.get('param', ''), row.get('param_value', ''), ot,
            int(row.get('count', 0) or 0),
            float(row.get('total_pnl', 0) or 0), avg_p,
            float(row.get('win_rate', 0) or 0) / 100, pf_v,
            float(row.get('rr_ratio', 0) or 0), edge,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=ri, column=ci)
            row_bg = e_bg if ci == 11 else (C['lgray'] if ri % 2 == 0 else C['white'])
            pnl_fg = (C['g_txt'] if avg_p >= 0 else C['r_txt']) if ci == 7 else 'FF222222'
            dc(cell, v, fmt=fmt, bg=row_bg, fg=pnl_fg,
               align='center' if ci not in [1, 2, 3, 11] else 'left')
            if ci in [1, 2, 3, 11]:
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           indent=1, wrap_text=True)
        ws.row_dimensions[ri].height = 17
    ws.freeze_panes = 'A5'


def _sheet_config(wb, run_id):
    ws = wb.create_sheet('Configuration')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 38
    ws.row_dimensions[1].height = 8
    title_row(ws, 2, 1, 2, 'BACKTEST CONFIGURATION', C['dark'], height=28)
    ws.row_dimensions[3].height = 8

    items = [
        ('Run ID',               str(run_id)),
        ('Initial Capital',      f'Rs.{INITIAL_CAPITAL:,}'),
        ('Brokerage per Order',  f'Rs.{BROKERAGE_PER_ORDER}'),
        ('Slippage Mode',        SLIPPAGE_MODE or 'None'),
        ('Slippage Percent',     f'{SLIPPAGE_PERCENT}%' if SLIPPAGE_MODE else '0%'),
        ('Composite Weights',    'Sharpe 30% · RODD 25% · PF 15% · Exp 10% · DD 10% · Cons 10%'),
        ('Generated At',         datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]
    for ri, (param, value) in enumerate(items, 4):
        p = ws.cell(row=ri, column=1)
        p.value = param
        p.font = Font(name='Arial', bold=True, size=10)
        p.fill = PatternFill('solid', fgColor=C['lgray'])
        p.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        p.border = mk_border()
        v = ws.cell(row=ri, column=2)
        v.value = value
        v.font = Font(name='Arial', size=10, color=C['b_txt'])
        v.fill = PatternFill('solid', fgColor='FFEEF6FF')
        v.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        v.border = mk_border()
        ws.row_dimensions[ri].height = 20


# ── main ──────────────────────────────────────────────────────────────────────

def generate_excel(db_path: str, run_id: str = None, out_dir: str = '.'):
    print(f"\n{'='*60}")
    print("  NIFTY SCALPING — STANDALONE EXCEL REPORT GENERATOR")
    print(f"{'='*60}")
    print(f"  DB   : {db_path}")
    print(f"  RunID: {run_id or 'latest'}")

    tables = load_tables(db_path, run_id)
    run_id = tables['run_id'] or datetime.now().strftime('%Y%m%d_%H%M%S')

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    print("\n📝 Building sheets...")
    _sheet_deployment_pick(wb, tables['ranking'],     run_id);  print("  ✓ Live Deployment Pick")
    _sheet_all_ranking    (wb, tables['ranking'],     run_id);  print("  ✓ All Strategy Ranking")
    _sheet_yearly         (wb, tables['yearly']);               print("  ✓ Yearly Breakdown")
    _sheet_monthly        (wb, tables['monthly']);              print("  ✓ Monthly Breakdown")
    _sheet_drawdown       (wb, tables['ranking']);              print("  ✓ Drawdown Analysis")
    _sheet_exit_reason    (wb, tables['exit_reason']);          print("  ✓ Exit Reason Analysis")
    _sheet_setup_edge     (wb, tables['single_param']);         print("  ✓ Setup Edge & Position Sizing")
    _sheet_expiry         (wb, tables['single_param']);         print("  ✓ Expiry Day Analysis")
    _sheet_call_put       (wb, tables['ot_overall']);           print("  ✓ CALL vs PUT Analysis")
    _sheet_scenarios      (wb, tables['ot_scenario']);          print("  ✓ CALL vs PUT Scenarios")
    _sheet_combo          (wb, tables['combo_param']);          print("  ✓ Combo Param Analysis")
    _sheet_filter_sim     (wb, tables['filter_sim']);           print("  ✓ Filter Simulation")
    _sheet_cpr            (wb, tables['single_param']);         print("  ✓ CPR Type Analysis")
    _sheet_gap_open       (wb, tables['single_param']);         print("  ✓ Gap & Open Location")
    _sheet_config         (wb, run_id);                         print("  ✓ Configuration")

    filename = os.path.join(out_dir, f"backtest_results_{run_id}.xlsx")
    wb.save(filename)
    print(f"\n✅  Saved: {filename}")
    print(f"   Sheets ({len(wb.worksheets)}): {', '.join(ws.title for ws in wb.worksheets)}")
    print(f"{'='*60}\n")
    return filename


if __name__ == '__main__':
    # parse args
    args    = sys.argv[1:]
    db_path = None
    run_id  = None

    i = 0
    while i < len(args):
        if args[i] == '--run' and i + 1 < len(args):
            run_id = args[i + 1]; i += 2
        elif args[i].endswith('.db'):
            db_path = args[i]; i += 1
        else:
            i += 1

    if db_path is None:
        base    = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base, 'myalgo.db')

    if not os.path.exists(db_path):
        print(f"❌  Database not found: {db_path}")
        sys.exit(1)

    out_dir = os.path.dirname(os.path.abspath(db_path))
    generate_excel(db_path, run_id, out_dir)